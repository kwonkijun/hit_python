import openpyxl

save_path = 'chapter3/스마트스토어.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path)

# data 시트 선택
ws = wb['data']

# (1) Range
ws['A1'] = '순번'
ws['B1'] = '제품명'
ws['C1'] = '가격'
ws['D1'] = '수량'
ws['E1'] = '합계'

# (2) Cell
ws.cell(row=2, column=1, value=1)
ws.cell(row=2, column=2, value='게이밍 마우스')
ws.cell(row=2, column=3, value=50000)
ws.cell(row=2, column=4, value=30)
ws.cell(row=2, column=5, value='=C2*D2')

# (3) 리스트
ws.append([2, '기계식 키보드', 120000, 15, '=C3*D3'])
ws.append([3, '마우스 패드', 20000, 5, '=C4*D4'])

# # 데이터 삭제
# del ws['A3']

# 엑셀 저장
wb.save(save_path)