import openpyxl

# 새로운 엑셀 파일 생성
total_wb = openpyxl.Workbook()

# 현재 활성화된 시트 선택
total_ws = total_wb.active

# 시트 이름 변경
total_ws.title = "data"

# 헤더 추가
total_ws.append(['순번', '제품명', '가격', '수량', '합계'])

# 데이터 파일
file_list = ['11번가', '스마트스토어', '쿠팡']

for file in file_list:
    wb = openpyxl.load_workbook(f'chapter3/{file}.xlsx', data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row = 2):
        data = []
        for cell in row:
            data.append(cell.value)
        total_ws.append(data)

for row in total_ws.iter_rows(min_row=2, max_col=1):
    for cell in row:
        cell.value = row[0].row - 1

i = 0
for cell in total_ws['A']:
    if i != 0:
        cell.value = i
    i = i + 1

total_wb.save('chapter3/total.xlsx')